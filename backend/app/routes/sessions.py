from __future__ import annotations

from io import BytesIO
from typing import List

import pandas as pd
import json

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from app.models.schemas import (
    ApplyOperationsRequest,
    ApplyOperationsResponse,
    BatchApplyRequest,
    CommitSummary,
    CreateEmptyWorkbookRequest,
    CreateSessionRequest,
    HistoryResponse,
    PreviewRequest,
    RollbackRequest,
)
from app.services import excel
from app.services.store import STORE


router = APIRouter()


@router.post("")
def create_session(payload: CreateSessionRequest):
    session = STORE.create_session(payload.name)
    return {"id": session.id, "name": session.name}


@router.post("/{session_id}/workbooks/empty")
def create_empty_workbook(session_id: str, payload: CreateEmptyWorkbookRequest):
    try:
        session = STORE.get_session(session_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="session_not_found")
    sheets = excel.create_empty(payload.sheet_name)
    STORE.add_workbook(session, payload.filename, sheets)
    return {"filename": payload.filename, "sheets": list(sheets.keys())}


@router.post("/{session_id}/workbooks/upload")
def upload_workbook(session_id: str, file: UploadFile = File(...)):
    try:
        session = STORE.get_session(session_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="session_not_found")
    content = file.file.read()
    sheets = excel.load_excel(content, file.filename)
    STORE.add_workbook(session, file.filename or "upload.xlsx", sheets)
    return {"filename": file.filename, "sheets": list(sheets.keys())}


@router.post("/{session_id}/workbooks/{filename}/preview")
def preview_workbook(session_id: str, filename: str, payload: PreviewRequest):
    try:
        session = STORE.get_session(session_id)
        workbook = STORE.get_workbook(session, filename)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    sheet_name = payload.sheet or (list(workbook.sheets.keys())[0] if workbook.sheets else "Sheet1")
    df = workbook.sheets.get(sheet_name)
    columns, rows = excel.preview(df, payload.limit)
    rules = [
        rule
        for rule in workbook.format_rules
        if rule.get("sheet") == sheet_name and rule.get("type") == "lt"
    ]
    return {
        "sheet": sheet_name,
        "columns": columns,
        "rows": rows,
        "rules": rules,
        "row_count": int(df.shape[0]) if df is not None else 0,
        "col_count": int(df.shape[1]) if df is not None else 0,
    }


@router.post("/{session_id}/workbooks/{filename}/operations", response_model=ApplyOperationsResponse)
def apply_operations(session_id: str, filename: str, payload: ApplyOperationsRequest):
    try:
        session = STORE.get_session(session_id)
        workbook = STORE.get_workbook(session, filename)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    ops = [op.model_dump(by_alias=True) for op in payload.operations]
    try:
    changed_sheets, analysis = excel.apply_operations(workbook.sheets, ops, workbook.format_rules)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    commit = STORE.commit(workbook, payload.message or "update", changed_sheets)
    return {
        "commit": CommitSummary(
            id=commit.id,
            message=commit.message,
            timestamp=commit.timestamp,
            changed_sheets=commit.changed_sheets,
        )
        ,
        "analysis": analysis,
    }


@router.post("/{session_id}/workbooks/{filename}/history", response_model=HistoryResponse)
def history(session_id: str, filename: str):
    try:
        session = STORE.get_session(session_id)
        workbook = STORE.get_workbook(session, filename)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    commits = [
        CommitSummary(
            id=commit.id,
            message=commit.message,
            timestamp=commit.timestamp,
            changed_sheets=commit.changed_sheets,
        )
        for commit in STORE.history(workbook)
    ]
    return {"commits": commits}


@router.post("/{session_id}/workbooks/{filename}/rollback", response_model=ApplyOperationsResponse)
def rollback(session_id: str, filename: str, payload: RollbackRequest):
    try:
        session = STORE.get_session(session_id)
        workbook = STORE.get_workbook(session, filename)
        commit = STORE.rollback(workbook, payload.commit_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return {
        "commit": CommitSummary(
            id=commit.id,
            message=commit.message,
            timestamp=commit.timestamp,
            changed_sheets=commit.changed_sheets,
        )
    }


@router.get("/{session_id}/workbooks/{filename}/export")
def export_workbook(session_id: str, filename: str, format: str = "xlsx"):
    try:
        session = STORE.get_session(session_id)
        workbook = STORE.get_workbook(session, filename)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    format = format.lower()
    if format not in {"xlsx", "csv"}:
        raise HTTPException(status_code=400, detail="invalid_format")
    if format == "csv":
        sheet_name = list(workbook.sheets.keys())[0] if workbook.sheets else "Sheet1"
        df = workbook.sheets.get(sheet_name, pd.DataFrame())
        content = df.to_csv(index=False).encode("utf-8")
        return StreamingResponse(BytesIO(content), media_type="text/csv")
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in workbook.sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    buffer.seek(0)
    if workbook.format_rules:
        wb = load_workbook(buffer)
        for rule in workbook.format_rules:
            sheet_name = rule.get("sheet")
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if not header_cells:
                continue
            try:
                col_idx = header_cells.index(rule.get("column")) + 1
            except ValueError:
                continue
            col_letter = get_column_letter(col_idx)
            if rule.get("type") == "number_format":
                number_format = rule.get("format") or "0.0"
                for cell in ws[f"{col_letter}2": f"{col_letter}{ws.max_row}"]:
                    for c in cell:
                        c.number_format = number_format
            elif rule.get("type") == "lt":
                threshold = rule.get("threshold")
                color = rule.get("color") or "red"
                if threshold is None:
                    continue
                fill = PatternFill(start_color="FFC7CE" if color == "red" else "FFFDE68A", end_color="FFC7CE" if color == "red" else "FFFDE68A", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator="lessThan", formula=[str(threshold)], fill=fill),
                )
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        buffer = out
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/{session_id}/batch")
def batch_apply(
    session_id: str,
    payload: str = Form(...),
    files: List[UploadFile] = File(...),
):
    try:
        session = STORE.get_session(session_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="session_not_found")
    try:
        payload_obj = BatchApplyRequest.model_validate(json.loads(payload))
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_payload")
    results = []
    for file in files:
        content = file.file.read()
        sheets = excel.load_excel(content, file.filename)
        workbook = STORE.add_workbook(session, file.filename or "upload.xlsx", sheets)
        ops = [op.model_dump(by_alias=True) for op in payload_obj.operations]
        changed_sheets, _analysis = excel.apply_operations(workbook.sheets, ops, workbook.format_rules)
        commit = STORE.commit(workbook, payload_obj.message or "batch", changed_sheets)
        results.append(
            {
                "filename": workbook.filename,
                "commit_id": commit.id,
                "changed_sheets": commit.changed_sheets,
            }
        )
    return {"results": results}
